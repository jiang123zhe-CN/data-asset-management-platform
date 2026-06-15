import io
import json
from datetime import datetime, timezone
from typing import BinaryIO

import openpyxl
from sqlalchemy.orm import Session

from app.models.field import Field, ImportRecord
from app.models.user import User

TEMPLATE_HEADERS = [
    "field_code", "name", "english_name", "data_type", "length",
    "precision", "table_name", "database_name", "business_domain",
    "sensitivity_level(自动)", "description", "business_rules",
]


def generate_template() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "字段导入模板"
    ws.append(TEMPLATE_HEADERS)
    ws.append(["F001", "客户名称", "customer_name", "VARCHAR", 100, None, "dim_customer", "ods", "客户域", "", "客户姓名", ""])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_excel(file: BinaryIO) -> list[dict]:
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [h.value for h in ws[1]]

    results = []
    for row_idx, row in enumerate(rows, start=2):
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, header in enumerate(["field_code", "name", "english_name", "data_type", "length",
                                     "precision", "table_name", "database_name", "business_domain",
                                     "sensitivity_level", "description", "business_rules"]):
            val = row[i] if i < len(row) else None
            row_dict[header] = str(val).strip() if val is not None else None
        row_dict["_row"] = row_idx
        results.append(row_dict)

    wb.close()
    return results


def import_fields(db: Session, file: BinaryIO, user: User, file_name: str, file_size: int) -> ImportRecord:
    rows = parse_excel(file)
    total = len(rows)
    errors = []
    success = 0
    new_field_ids = []

    import_record = ImportRecord(
        user_id=user.id,
        file_name=file_name,
        file_size=file_size,
        total_rows=total,
        success_rows=0,
        failed_rows=0,
        status="processing",
    )
    db.add(import_record)
    db.flush()

    for row in rows:
        rownum = row.pop("_row")
        row_errors = []

        if not row.get("field_code"):
            row_errors.append({"row": rownum, "field": "field_code", "message": "字段编码不能为空"})
        if not row.get("name"):
            row_errors.append({"row": rownum, "field": "name", "message": "字段名称不能为空"})
        if not row.get("data_type"):
            row_errors.append({"row": rownum, "field": "data_type", "message": "数据类型不能为空"})
        if not row.get("table_name"):
            row_errors.append({"row": rownum, "field": "table_name", "message": "表名不能为空"})

        # sensitivity_level is optional - system auto-classifies
        sl = row.get("sensitivity_level") or "L2"

        existing = db.query(Field).filter(Field.field_code == row["field_code"]).first()
        if existing:
            row_errors.append({"row": rownum, "field": "field_code", "message": "字段编码已存在"})

        length_val = None
        if row.get("length"):
            try:
                length_val = int(row["length"])
            except ValueError:
                row_errors.append({"row": rownum, "field": "length", "message": "长度必须为整数"})

        precision_val = None
        if row.get("precision"):
            try:
                precision_val = int(row["precision"])
            except ValueError:
                row_errors.append({"row": rownum, "field": "precision", "message": "精度必须为整数"})

        if row_errors:
            errors.extend(row_errors)
            continue

        field = Field(
            field_code=row["field_code"],
            name=row["name"],
            english_name=row.get("english_name"),
            data_type=row["data_type"],
            length=length_val,
            precision=precision_val,
            table_name=row["table_name"],
            database_name=row.get("database_name"),
            business_domain=row.get("business_domain"),
            sensitivity_level=sl,
            description=row.get("description"),
            business_rules=row.get("business_rules"),
            source="excel_import",
            import_batch_id=import_record.id,
            created_by=user.id,
        )
        db.add(field)
        db.flush()
        new_field_ids.append(field.id)
        success += 1

    import_record.success_rows = success
    import_record.failed_rows = total - success
    import_record.status = "completed" if success == total else ("partial" if success > 0 else "failed")
    import_record.error_details = json.dumps(errors, ensure_ascii=False) if errors else None
    db.commit()
    db.refresh(import_record)

    # Auto-classify all newly imported fields via compliance engine
    if new_field_ids:
        from app.services.compliance_engine import ComplianceEngine
        comp = ComplianceEngine(db)
        comp.classify_fields(new_field_ids)
        db.commit()

    return import_record


def export_mappings_to_excel(mappings: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产映射数据"

    headers = [
        "映射ID", "目录路径", "目录编码", "字段编码", "字段名称",
        "数据类型", "来源表", "映射来源", "置信度", "创建时间",
    ]
    ws.append(headers)

    for m in mappings:
        ws.append([
            m.get("id"),
            m.get("directory_path", ""),
            m.get("directory_code", ""),
            m.get("field_code", ""),
            m.get("field_name", ""),
            m.get("field_data_type", ""),
            m.get("field_table", ""),
            "AI建议" if m.get("mapping_source") == "ai_suggested" else "手动映射",
            m.get("confidence"),
            m.get("created_at"),
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_compliance_inventory(db: Session) -> io.BytesIO:
    """导出合规分类分级清单 — 符合指南§6.4标准格式。"""
    from app.models.finance_category import FinanceDataCategory

    # ① 加载所有活跃字段 + 金融分类 + 层级信息
    fields = (
        db.query(Field)
        .filter(Field.status == "active", Field.finance_category_id.isnot(None))
        .order_by(Field.finance_data_level.desc(), Field.table_name, Field.field_code)
        .all()
    )

    all_cats = (
        db.query(FinanceDataCategory)
        .filter(FinanceDataCategory.is_active == True)
        .all()
    )
    cat_by_id = {c.id: c for c in all_cats}
    id_to_parent = {c.id: c for c in all_cats}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "金融数据分类分级清单"

    # ② 表头（指南 §6.4 格式）
    headers = [
        "序号", "字段编码", "字段名称", "表名", "数据库",
        "一级分类", "二级分类", "三级分类", "分类编码",
        "数据描述", "数据示例", "参考最低级别",
        "实际合规级别", "级别是否升级", "分级理由",
        "业务域", "数据来源", "最后更新日期",
    ]
    ws.append(headers)

    # ③ 填充数据行
    LEVEL_LABELS = {"core": "核心数据", "important": "重要数据", "sensitive": "敏感一般数据", "normal": "常规一般数据"}

    for idx, f in enumerate(fields, start=1):
        cat = cat_by_id.get(f.finance_category_id) if f.finance_category_id else None
        cat_code = cat.code if cat else None
        cat_name = cat.name if cat else None

        # 构建三级路径
        l1_name, l2_name, l3_name = "", "", ""
        if cat:
            path = [cat]
            pid = cat.parent_id
            while pid and pid in id_to_parent:
                path.append(id_to_parent[pid])
                pid = id_to_parent[pid].parent_id
            path.reverse()
            if len(path) >= 1:
                l1_name = path[0].name
            if len(path) >= 2:
                l2_name = path[1].name
            l3_name = cat.name

        # 级别比较
        ref_level = cat.ref_min_level if cat else "normal"
        actual_level = f.finance_data_level or "normal"
        LEVEL_ORDER = {"core": 4, "important": 3, "sensitive": 2, "normal": 1}
        upgraded = LEVEL_ORDER.get(actual_level, 0) > LEVEL_ORDER.get(ref_level, 0)

        # 分级理由
        if upgraded:
            upgrade_reason = f"参考最低级别为{LEVEL_LABELS.get(ref_level, ref_level)}，因数据特征触发矩阵升级至{LEVEL_LABELS.get(actual_level, actual_level)}"
        else:
            upgrade_reason = f"与参考最低级别一致（{LEVEL_LABELS.get(ref_level, ref_level)}）"

        ws.append([
            idx,
            f.field_code, f.name, f.table_name, f.database_name,
            l1_name, l2_name, l3_name, cat_code,
            cat.appendix_desc if cat else None,
            cat.appendix_example if cat else None,
            LEVEL_LABELS.get(ref_level, ref_level),
            LEVEL_LABELS.get(actual_level, actual_level),
            "是" if upgraded else "否",
            upgrade_reason,
            f.business_domain, f.source,
            f.updated_at.strftime("%Y-%m-%d %H:%M") if f.updated_at else "",
        ])

    # ④ 调整列宽
    col_widths = {
        "A": 6, "B": 14, "C": 16, "D": 16, "E": 14,
        "F": 14, "G": 14, "H": 14, "I": 24,
        "J": 30, "K": 30, "L": 14,
        "M": 14, "N": 12, "O": 40,
        "P": 12, "Q": 12, "R": 18,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ⑤ 表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ⑥ 冻结首行
    ws.freeze_panes = "A2"

    # ⑦ 添加摘要 sheet
    ws2 = wb.create_sheet("统计摘要")
    total = len(fields)
    all_fields = db.query(Field).filter(Field.status == "active").count()
    coverage = f"{total / all_fields * 100:.1f}%" if all_fields > 0 else "0%"
    level_counts = {}
    for f in fields:
        lv = f.finance_data_level or "unclassified"
        level_counts[lv] = level_counts.get(lv, 0) + 1

    summary_data = [
        ["合规分类分级清单 — 统计摘要"],
        [""],
        ["导出时间", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ["标准依据", "《金融信息服务数据分类分级指南》国信办通字〔2026〕2号"],
        ["活跃字段总数", all_fields],
        ["已分类字段数", total],
        ["分类覆盖率", coverage],
        [""],
        ["级别分布", ""],
    ]
    for lv in ["core", "important", "sensitive", "normal"]:
        summary_data.append([f"  {LEVEL_LABELS.get(lv, lv)}", level_counts.get(lv, 0)])
    summary_data.append([" 未分类", level_counts.get("unclassified", 0)])

    for row in summary_data:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_fields_to_excel(fields: list[Field]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据字段"
    ws.append(TEMPLATE_HEADERS + ["status", "is_anomaly"])

    for f in fields:
        ws.append([
            f.field_code, f.name, f.english_name, f.data_type, f.length,
            f.precision, f.table_name, f.database_name, f.business_domain,
            f.sensitivity_level, f.description, f.business_rules,
            f.status, str(f.is_anomaly),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_tagging_results(db: Session) -> io.BytesIO:
    """导出数据打标结果 — 全部活跃字段的合规分类+级别+方法+置信度。"""
    from app.models.finance_category import FinanceDataCategory

    fields = (
        db.query(Field)
        .filter(Field.status == "active")
        .order_by(Field.finance_data_level.desc().nullslast(), Field.table_name, Field.field_code)
        .all()
    )

    # ① 加载分类映射 + 层级路径
    fin_cat_ids = [f.finance_category_id for f in fields if f.finance_category_id]
    all_fin = (
        db.query(FinanceDataCategory)
        .filter(FinanceDataCategory.is_active == True)
        .all()
    )
    cat_by_id = {c.id: c for c in all_fin}
    id_to_parent = {c.id: c for c in all_fin}
    fin_path_map: dict[int, str] = {}
    for cid in fin_cat_ids:
        c = cat_by_id.get(cid)
        if c:
            parts = [c.name]
            pid = c.parent_id
            while pid and pid in id_to_parent:
                parts.append(id_to_parent[pid].name)
                pid = id_to_parent[pid].parent_id
            fin_path_map[cid] = " > ".join(reversed(parts))

    LEVEL_LABELS = {"core": "核心数据", "important": "重要数据", "sensitive": "敏感一般数据", "normal": "常规一般数据"}
    METHOD_LABELS = {"compliance_matrix": "合规矩阵", "rule_engine": "规则引擎", "ai": "AI辅助", "manual": "人工", "hybrid": "混合"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据打标结果"

    headers = [
        "序号", "字段编码", "字段名称", "英文名", "数据类型",
        "表名", "数据库", "业务域",
        "金融合规分类（完整路径）", "分类编码",
        "合规数据级别", "打标方法", "置信度(%)",
        "最后打标时间", "是否异常", "来源",
    ]
    ws.append(headers)

    for idx, f in enumerate(fields, start=1):
        cat = cat_by_id.get(f.finance_category_id) if f.finance_category_id else None
        cat_path = fin_path_map.get(f.finance_category_id, "") if f.finance_category_id else ""
        cat_code = cat.code if cat else ""
        actual_level = f.finance_data_level
        level_label = LEVEL_LABELS.get(actual_level, "") if actual_level else ""
        method_label = METHOD_LABELS.get(f.tagging_method, f.tagging_method or "")
        confidence_pct = round((f.tagging_confidence or 0) * 100)

        ws.append([
            idx,
            f.field_code, f.name, f.english_name, f.data_type,
            f.table_name, f.database_name, f.business_domain,
            cat_path, cat_code,
            level_label, method_label, confidence_pct,
            f.last_tagged_at.strftime("%Y-%m-%d %H:%M") if f.last_tagged_at else "",
            "异常" if f.is_anomaly else "正常",
            "Excel导入" if f.source == "excel_import" else "手动",
        ])

    # ④ 调整列宽
    col_widths = {
        "A": 6, "B": 14, "C": 16, "D": 14, "E": 12,
        "F": 16, "G": 14, "H": 12,
        "I": 40, "J": 28,
        "K": 14, "L": 12, "M": 10,
        "N": 18, "O": 10, "P": 12,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # ⑥ 摘要 sheet
    ws2 = wb.create_sheet("统计摘要")
    total = len(fields)
    classified = sum(1 for f in fields if f.finance_data_level)
    level_counts: dict[str, int] = {}
    for f in fields:
        lv = f.finance_data_level or "unclassified"
        level_counts[lv] = level_counts.get(lv, 0) + 1
    method_counts: dict[str, int] = {}
    for f in fields:
        m = f.tagging_method or "unset"
        method_counts[m] = method_counts.get(m, 0) + 1

    summary_rows = [
        ["数据打标结果 — 统计摘要"],
        [""],
        ["导出时间", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ["活跃字段总数", total],
        ["已打标字段数", classified],
        ["打标覆盖率", f"{classified / total * 100:.1f}%" if total > 0 else "0%"],
        [""],
        ["合规级别分布", ""],
    ]
    for lv in ["core", "important", "sensitive", "normal"]:
        summary_rows.append([f"  {LEVEL_LABELS.get(lv, lv)}", level_counts.get(lv, 0)])
    summary_rows.append([" 未分类", level_counts.get("unclassified", 0)])
    summary_rows.append([""])
    summary_rows.append(["打标方法分布", ""])
    for m, label in METHOD_LABELS.items():
        c = method_counts.get(m, 0)
        if c > 0:
            summary_rows.append([f"  {label}", c])

    for row in summary_rows:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
